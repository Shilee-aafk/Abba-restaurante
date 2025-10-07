"""
Views for the Restaurante ABBA application.

This module contains all the view functions organized by functionality:
- Authentication views (register, home)
- Waiter views (select_table, menu, send_order, toggle_table_availability)
- Kitchen views (kitchen_queue, kitchen_queue_data, update_order_status)
- Admin views (admin_users, audit_log)
- Reception views (reception, download_daily_report)
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, F
from django.contrib.auth.models import User
from django.contrib.auth import login
from django.views.decorators.http import require_GET
from .models import Table, MenuItem, Order, OrderItem, UserProfile, AuditLog, RegistrationPIN
import json
import random
import string
from openpyxl import Workbook
from datetime import date
from collections import defaultdict


# Authentication Views

def register(request):
    """
    Handle user registration with PIN-based role assignment.

    Allows new users to register using a generated PIN that determines their role.
    After successful registration, automatically logs in the user and redirects to home.
    """
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        pin = request.POST.get('pin')

        if not username or not password or not pin:
            messages.error(request, 'Todos los campos son requeridos')
            return render(request, 'restaurant/register.html')

        try:
            pin_obj = RegistrationPIN.objects.get(pin=pin, uses__lt=2)
        except RegistrationPIN.DoesNotExist:
            messages.error(request, 'PIN inválido o agotado')
            return render(request, 'restaurant/register.html')

        if User.objects.filter(username=username).exists():
            messages.error(request, 'Usuario ya existe')
            return render(request, 'restaurant/register.html')

        user = User.objects.create_user(username=username, password=password)

        # Assign role based on PIN
        role_mapping = {
            'garzon': 'garzon',
            'cocinero': 'cocinero',
            'admin': 'admin',
            'recepcion': 'recepcion'
        }
        user.userprofile.role = role_mapping.get(pin_obj.role, 'garzon')
        user.userprofile.save()

        pin_obj.uses += 1
        pin_obj.save()

        # Log audit
        AuditLog.objects.create(
            user=user,
            action='Registro de usuario',
            details=f'Usuario {username} registrado con rol {user.userprofile.role}'
        )

        messages.success(request, 'Usuario registrado exitosamente')
        login(request, user)
        return redirect('home')

    return render(request, 'restaurant/register.html')


@login_required
def home(request):
    """
    Main dashboard view that redirects users to their role-specific page.

    Ensures UserProfile exists for the user and redirects based on their role.
    Superusers are redirected to Django admin.
    """
    # Ensure UserProfile exists
    UserProfile.objects.get_or_create(user=request.user, defaults={'role': 'garzon'})

    # If superuser, redirect to Django admin
    if request.user.is_superuser:
        return redirect('/admin/')

    role_redirects = {
        'garzon': 'select_table',
        'cocinero': 'kitchen_queue',
        'admin': 'admin_users',
        'recepcion': 'reception'
    }

    redirect_url = role_redirects.get(request.user.userprofile.role)
    if redirect_url:
        return redirect(redirect_url)
    else:
        return render(request, 'restaurant/home.html', {'message': 'Rol no reconocido'})


# Waiter Views

@login_required
def select_table(request):
    """
    Display available tables for order placement.

    Accessible by waiters and admins. Shows all tables with their availability status.
    """
    if request.user.userprofile.role not in ['garzon', 'admin']:
        return redirect('home')
    tables = Table.objects.all()
    return render(request, 'restaurant/select_table.html', {'tables': tables})


@login_required
def menu(request, table_id):
    """
    Display menu for a specific table.

    Shows available menu items that can be ordered for the selected table.
    """
    if request.user.userprofile.role not in ['garzon', 'admin']:
        return redirect('home')
    table = get_object_or_404(Table, id=table_id)
    products = MenuItem.objects.filter(available=True)
    return render(request, 'restaurant/menu.html', {'table': table, 'products': products})


@login_required
def send_order(request, table_id):
    """
    Process order submission for a table.

    Creates a new order with selected items, marks table as occupied,
    and logs the action in audit log.
    """
    if request.user.userprofile.role not in ['garzon', 'admin']:
        return JsonResponse({'error': 'No autorizado'}, status=403)

    if request.method == 'POST':
        items = []
        notes = request.POST.get('order_notes', '')

        # Extract items from POST data
        for key in request.POST:
            if key.startswith('item_id_'):
                suffix = key[len('item_id_'):]
                try:
                    item_id = int(request.POST.get(f'item_id_{suffix}'))
                    quantity = int(request.POST.get(f'quantity_{suffix}', 1))
                    if quantity <= 0:
                        continue
                    item_notes = request.POST.get(f'notes_{suffix}', '')
                    items.append({'id': item_id, 'quantity': quantity, 'notes': item_notes})
                except (ValueError, TypeError):
                    continue

        if not items:
            return JsonResponse({'error': 'No hay ítems en el pedido'}, status=400)

        table = get_object_or_404(Table, id=table_id)
        order = Order.objects.create(table=table, waiter=request.user, notes=notes)

        for item in items:
            menu_item = get_object_or_404(MenuItem, id=item['id'])
            OrderItem.objects.create(
                order=order,
                menu_item=menu_item,
                quantity=item['quantity'],
                notes=item.get('notes', '')
            )

        # Mark table as occupied
        table.is_available = False
        table.save()

        # Log audit
        AuditLog.objects.create(
            user=request.user,
            action='Crear pedido',
            details=f'Pedido {order.id} para mesa {table.number}'
        )

        return redirect('select_table')

    return JsonResponse({'error': 'Método no permitido'}, status=405)


@login_required
def toggle_table_availability(request, table_id):
    """
    Toggle table availability status.

    Allows waiters and admins to mark tables as available or occupied.
    """
    if request.user.userprofile.role not in ['garzon', 'admin']:
        return JsonResponse({'error': 'No autorizado'}, status=403)

    if request.method == 'POST':
        table = get_object_or_404(Table, id=table_id)
        table.is_available = not table.is_available
        table.save()

        # Log audit
        status_text = 'disponible' if table.is_available else 'ocupada'
        AuditLog.objects.create(
            user=request.user,
            action=f'Marcar mesa como {status_text}',
            details=f'Mesa {table.number} marcada como {status_text}'
        )

        return JsonResponse({'success': True, 'is_available': table.is_available})

    return JsonResponse({'error': 'Método no permitido'}, status=405)


# Kitchen Views

@login_required
def kitchen_queue(request):
    """
    Display kitchen queue with pending orders.

    Shows orders that are not taken or in preparation, grouped by items.
    """
    if request.user.userprofile.role not in ['cocinero', 'admin']:
        return redirect('home')

    orders = Order.objects.filter(status__in=['not_taken', 'preparing'])\
                         .prefetch_related('items')\
                         .order_by('created_at')

    # Group items by product and notes for each order
    for order in orders:
        grouped_items = defaultdict(lambda: {'menu_item': None, 'quantity': 0, 'notes': ''})
        for item in order.items.all():
            key = (item.menu_item.id, item.notes.strip())
            if grouped_items[key]['menu_item'] is None:
                grouped_items[key]['menu_item'] = item.menu_item
                grouped_items[key]['notes'] = item.notes.strip()
            grouped_items[key]['quantity'] += item.quantity
        order.grouped_items = list(grouped_items.values())

    return render(request, 'restaurant/kitchen_queue.html', {'orders': orders})


@login_required
@require_GET
def kitchen_queue_data(request):
    """
    API endpoint for kitchen queue data.

    Returns JSON data for AJAX updates of the kitchen queue.
    """
    if request.user.userprofile.role not in ['cocinero', 'admin']:
        return JsonResponse({'error': 'No autorizado'}, status=403)

    orders = Order.objects.filter(status__in=['not_taken', 'preparing'])\
                         .prefetch_related('items')\
                         .order_by('created_at')

    data = []
    for order in orders:
        grouped_items = defaultdict(lambda: {'menu_item_name': '', 'quantity': 0, 'notes': ''})
        for item in order.items.all():
            key = (item.menu_item.id, item.notes.strip())
            if grouped_items[key]['menu_item_name'] == '':
                grouped_items[key]['menu_item_name'] = item.menu_item.name
                grouped_items[key]['notes'] = item.notes.strip()
            grouped_items[key]['quantity'] += item.quantity
        items = list(grouped_items.values())
        data.append({
            'id': order.id,
            'table_number': order.table.number,
            'created_at_date': order.created_at.strftime('%d/%m/%Y'),
            'created_at_time': order.created_at.strftime('%H:%M'),
            'status': order.status,
            'status_display': order.get_status_display(),
            'notes': order.notes,
            'items': items,
        })
    return JsonResponse({'orders': data})


@login_required
def update_order_status(request, order_id):
    """
    Update order status (preparing or ready).

    Allows kitchen staff to change order status and logs the action.
    """
    if request.user.userprofile.role not in ['cocinero', 'admin']:
        return JsonResponse({'error': 'No autorizado'}, status=403)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            new_status = data.get('status')
            if new_status not in ['preparing', 'ready']:
                return JsonResponse({'error': 'Estado inválido'}, status=400)

            order = get_object_or_404(Order, id=order_id)
            order.status = new_status
            order.save()

            # Log audit
            AuditLog.objects.create(
                user=request.user,
                action=f'Cambiar estado pedido a {new_status}',
                details=f'Pedido {order.id}'
            )

            return JsonResponse({'success': True})
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Datos JSON inválidos'}, status=400)
        except Exception as e:
            return JsonResponse({'error': 'Error interno del servidor'}, status=500)

    return JsonResponse({'error': 'Método no permitido'}, status=405)


# Admin Views

@login_required
def admin_users(request):
    """
    Admin interface for user management.

    Displays users and PINs, allows generating new registration PINs.
    """
    if request.user.userprofile.role != 'admin':
        return redirect('home')

    users = UserProfile.objects.all()
    pins = RegistrationPIN.objects.all().order_by('-created_at')

    if request.method == 'POST':
        role = request.POST.get('role')
        if role in ['garzon', 'cocinero', 'admin', 'recepcion']:
            pin = ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
            RegistrationPIN.objects.create(pin=pin, role=role, created_by=request.user)
            messages.success(request, f'PIN generado: {pin} para rol {role}')

            # Log audit
            AuditLog.objects.create(
                user=request.user,
                action='Generar PIN de registro',
                details=f'PIN {pin} para rol {role}'
            )
        else:
            messages.error(request, 'Rol inválido')
        return redirect('admin_users')

    return render(request, 'restaurant/admin_users.html', {'users': users, 'pins': pins})


@login_required
def audit_log(request):
    """
    Display audit log for administrative review.

    Shows all logged actions in reverse chronological order.
    """
    if request.user.userprofile.role != 'admin':
        return redirect('home')
    logs = AuditLog.objects.all().order_by('-timestamp')
    return render(request, 'restaurant/audit_log.html', {'logs': logs})


# Reception Views

@login_required
def reception(request):
    """
    Reception dashboard with daily sales summary.

    Shows today's orders and total sales for reception staff.
    """
    if request.user.userprofile.role != 'recepcion':
        return redirect('home')

    today = date.today()
    orders = Order.objects.filter(created_at__date=today)\
                         .prefetch_related('items__menu_item')\
                         .annotate(total=Sum(F('items__quantity') * F('items__menu_item__price')))

    total_general = sum(order.total or 0 for order in orders)
    return render(request, 'restaurant/reception.html', {'orders': orders, 'total_general': total_general})


@login_required
def download_daily_report(request):
    """
    Generate and download daily sales report as Excel file.

    Creates an Excel spreadsheet with today's orders and totals.
    """
    if request.user.userprofile.role != 'recepcion':
        return redirect('home')

    today = date.today()
    orders = Order.objects.filter(created_at__date=today)

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Diario"

    # Headers
    headers = ['ID Pedido', 'Mesa', 'Garzón', 'Hora', 'Estado', 'Total']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    row = 2
    total_general = 0
    for order in orders:
        total_order = sum(item.quantity * item.menu_item.price for item in order.items.all())
        total_general += total_order

        ws.cell(row=row, column=1, value=order.id)
        ws.cell(row=row, column=2, value=order.table.number)
        ws.cell(row=row, column=3, value=order.waiter.username)
        ws.cell(row=row, column=4, value=order.created_at.strftime('%H:%M'))
        ws.cell(row=row, column=5, value=order.get_status_display())
        ws.cell(row=row, column=6, value=float(total_order))
        row += 1

    # Total general
    ws.cell(row=row, column=5, value='Total General')
    ws.cell(row=row, column=6, value=float(total_general))

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=reporte_diario_{today}.xlsx'
    wb.save(response)
    return response
